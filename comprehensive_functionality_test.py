#!/usr/bin/env python3
"""
Comprehensive Functionality Test Suite
Tests all features of the Career Intelligence System
"""

import requests
import json
import time
import os
from pathlib import Path

# Configuration
API_BASE_URL = "http://localhost:8000/api/v1"
FRONTEND_URL = "http://localhost:3000"

# Test data
test_user_email = f"test_user_{int(time.time())}@example.com"
test_user_password = "TestPassword123!"
test_user_name = "Test User"

# Store test results
results = {
    "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
    "tests": [],
    "summary": {}
}

def log_test(name, status, details=""):
    """Log test result"""
    result = {
        "name": name,
        "status": status,
        "details": details,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
    }
    results["tests"].append(result)
    status_icon = "✅" if status == "PASS" else "❌"
    print(f"{status_icon} {name}: {details}")

def test_system_health():
    """Test 1: System Health Check"""
    try:
        response = requests.get(f"{API_BASE_URL}/system/status", timeout=5)
        if response.status_code == 200:
            data = response.json()
            log_test("System Health", "PASS", f"Status: {data.get('status')}")
            return True
        else:
            log_test("System Health", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("System Health", "FAIL", str(e))
        return False

def test_user_registration():
    """Test 2: User Registration"""
    try:
        payload = {
            "email": test_user_email,
            "password": test_user_password,
            "name": test_user_name
        }
        response = requests.post(f"{API_BASE_URL}/auth/register", json=payload, timeout=5)
        if response.status_code in [200, 201]:
            log_test("User Registration", "PASS", f"User: {test_user_email}")
            return True
        else:
            log_test("User Registration", "FAIL", f"Status: {response.status_code}, Response: {response.text}")
            return False
    except Exception as e:
        log_test("User Registration", "FAIL", str(e))
        return False

def test_user_login():
    """Test 3: User Login"""
    try:
        payload = {
            "email": test_user_email,
            "password": test_user_password
        }
        response = requests.post(f"{API_BASE_URL}/auth/login", json=payload, timeout=5)
        if response.status_code == 200:
            token = response.json().get("access_token")
            log_test("User Login", "PASS", f"Token generated")
            return token
        else:
            log_test("User Login", "FAIL", f"Status: {response.status_code}")
            return None
    except Exception as e:
        log_test("User Login", "FAIL", str(e))
        return None

def test_profile_creation(token):
    """Test 4: Profile Creation"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        payload = {
            "education_level": "Bachelor's Degree",
            "experience_years": 2.5,
            "skills": "Python, React, SQL, AWS",
            "interests": "Web Development, Cloud Computing",
            "bio": "Passionate developer with interest in full-stack development"
        }
        response = requests.post(f"{API_BASE_URL}/students/me", json=payload, headers=headers, timeout=5)
        if response.status_code in [200, 201]:
            log_test("Profile Creation", "PASS", "Profile created successfully")
            return True
        else:
            log_test("Profile Creation", "FAIL", f"Status: {response.status_code}, Response: {response.text}")
            return False
    except Exception as e:
        log_test("Profile Creation", "FAIL", str(e))
        return False

def test_profile_retrieval(token):
    """Test 5: Profile Retrieval"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        response = requests.get(f"{API_BASE_URL}/students/me", headers=headers, timeout=5)
        if response.status_code == 200:
            profile = response.json()
            log_test("Profile Retrieval", "PASS", f"Education: {profile.get('education_level')}")
            return True
        else:
            log_test("Profile Retrieval", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("Profile Retrieval", "FAIL", str(e))
        return False

def test_career_score(token):
    """Test 6: Career Readiness Score Calculation"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        response = requests.get(f"{API_BASE_URL}/career/score", headers=headers, timeout=5)
        if response.status_code == 200:
            score_data = response.json()
            score = score_data.get("score", 0)
            log_test("Career Score", "PASS", f"Score: {score}/100")
            return True
        else:
            log_test("Career Score", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("Career Score", "FAIL", str(e))
        return False

def test_job_recommendations(token):
    """Test 7: Job Recommendations"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        response = requests.get(f"{API_BASE_URL}/career/recommendations", headers=headers, timeout=5)
        if response.status_code == 200:
            recommendations = response.json()
            count = len(recommendations.get("recommendations", []))
            log_test("Job Recommendations", "PASS", f"Found {count} recommendations")
            return True
        else:
            log_test("Job Recommendations", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("Job Recommendations", "FAIL", str(e))
        return False

def test_ai_recommendations(token):
    """Test 8: AI Recommendations"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        response = requests.get(f"{API_BASE_URL}/career/ai-recommendations", headers=headers, timeout=5)
        if response.status_code == 200:
            log_test("AI Recommendations", "PASS", "AI recommendations generated")
            return True
        else:
            log_test("AI Recommendations", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("AI Recommendations", "FAIL", str(e))
        return False

def test_document_upload(token):
    """Test 9: Document Upload"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        
        # Create a test file
        test_file_path = "test_document.txt"
        with open(test_file_path, "w") as f:
            f.write("This is a test certificate with Python, React, and AWS skills mentioned.")
        
        with open(test_file_path, "rb") as f:
            files = {"file": (test_file_path, f, "text/plain")}
            response = requests.post(f"{API_BASE_URL}/documents/upload", files=files, headers=headers, timeout=10)
        
        # Clean up
        os.remove(test_file_path)
        
        if response.status_code in [200, 201]:
            log_test("Document Upload", "PASS", "Document uploaded successfully")
            return response.json().get("id")
        else:
            log_test("Document Upload", "FAIL", f"Status: {response.status_code}")
            return None
    except Exception as e:
        log_test("Document Upload", "FAIL", str(e))
        return None

def test_ocr_processing(token, doc_id):
    """Test 10: OCR Processing"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        response = requests.post(f"{API_BASE_URL}/documents/{doc_id}/ocr", headers=headers, timeout=30)
        if response.status_code == 200:
            doc_data = response.json()
            ocr_text = doc_data.get("ocr_text", "")
            confidence = doc_data.get("ocr_confidence", 0)
            log_test("OCR Processing", "PASS", f"OCR confidence: {confidence:.2f}")
            return True
        else:
            log_test("OCR Processing", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("OCR Processing", "FAIL", str(e))
        return False

def test_skill_extraction(token, doc_id):
    """Test 11: Skill Extraction"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        payload = {"document_ids": [doc_id]}
        response = requests.post(f"{API_BASE_URL}/documents/extract-skills", json=payload, headers=headers, timeout=10)
        if response.status_code == 200:
            skills_data = response.json()
            skills = skills_data.get("skills", [])
            log_test("Skill Extraction", "PASS", f"Extracted {len(skills)} skills")
            return True
        else:
            log_test("Skill Extraction", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("Skill Extraction", "FAIL", str(e))
        return False

def test_report_generation(token):
    """Test 12: Report Generation"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        response = requests.post(f"{API_BASE_URL}/reports/generate", headers=headers, timeout=30)
        if response.status_code in [200, 201]:
            log_test("Report Generation", "PASS", "Report generated successfully")
            return True
        else:
            log_test("Report Generation", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("Report Generation", "FAIL", str(e))
        return False

def test_journey_status(token):
    """Test 13: Journey Status"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        response = requests.get(f"{API_BASE_URL}/journey/status", headers=headers, timeout=5)
        if response.status_code == 200:
            journey_data = response.json()
            stage = journey_data.get("stage", 0)
            completion = journey_data.get("completion_percentage", 0)
            log_test("Journey Status", "PASS", f"Stage: {stage}, Completion: {completion}%")
            return True
        else:
            log_test("Journey Status", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("Journey Status", "FAIL", str(e))
        return False

def test_knowledge_base_upload(token):
    """Test 14: Knowledge Base Upload"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        
        # Create a test Excel file
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Role'
            ws['B1'] = 'Technical Skills'
            ws['A2'] = 'Python Developer'
            ws['B2'] = 'Python, Django, PostgreSQL'
            
            test_kb_path = "test_kb.xlsx"
            wb.save(test_kb_path)
            
            with open(test_kb_path, "rb") as f:
                files = {"file": (test_kb_path, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                response = requests.post(f"{API_BASE_URL}/kb/upload", files=files, headers=headers, timeout=30)
            
            os.remove(test_kb_path)
            
            if response.status_code in [200, 201]:
                log_test("KB Upload", "PASS", "Knowledge base uploaded successfully")
                return True
            else:
                log_test("KB Upload", "FAIL", f"Status: {response.status_code}")
                return False
        except ImportError:
            log_test("KB Upload", "SKIP", "openpyxl not available")
            return False
    except Exception as e:
        log_test("KB Upload", "FAIL", str(e))
        return False

def test_kb_search(token):
    """Test 15: Knowledge Base Search"""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        payload = {"query": "Python developer"}
        response = requests.post(f"{API_BASE_URL}/kb/search", json=payload, headers=headers, timeout=10)
        if response.status_code == 200:
            results = response.json()
            count = len(results.get("results", []))
            log_test("KB Search", "PASS", f"Found {count} results")
            return True
        else:
            log_test("KB Search", "FAIL", f"Status: {response.status_code}")
            return False
    except Exception as e:
        log_test("KB Search", "FAIL", str(e))
        return False

def run_all_tests():
    """Run all tests"""
    print("\n" + "="*80)
    print("COMPREHENSIVE FUNCTIONALITY TEST SUITE")
    print("="*80 + "\n")
    
    # Test 1: System Health
    if not test_system_health():
        print("\n❌ System is not healthy. Stopping tests.")
        return
    
    # Test 2: Registration
    if not test_user_registration():
        print("\n❌ Registration failed. Stopping tests.")
        return
    
    # Test 3: Login
    token = test_user_login()
    if not token:
        print("\n❌ Login failed. Stopping tests.")
        return
    
    # Test 4: Profile Creation
    test_profile_creation(token)
    
    # Test 5: Profile Retrieval
    test_profile_retrieval(token)
    
    # Test 6: Career Score
    test_career_score(token)
    
    # Test 7: Job Recommendations
    test_job_recommendations(token)
    
    # Test 8: AI Recommendations
    test_ai_recommendations(token)
    
    # Test 9: Document Upload
    doc_id = test_document_upload(token)
    
    # Test 10: OCR Processing
    if doc_id:
        test_ocr_processing(token, doc_id)
        
        # Test 11: Skill Extraction
        test_skill_extraction(token, doc_id)
    
    # Test 12: Report Generation
    test_report_generation(token)
    
    # Test 13: Journey Status
    test_journey_status(token)
    
    # Test 14: Knowledge Base Upload
    test_knowledge_base_upload(token)
    
    # Test 15: Knowledge Base Search
    test_kb_search(token)
    
    # Print summary
    print("\n" + "="*80)
    print("TEST SUMMARY")
    print("="*80)
    
    passed = sum(1 for t in results["tests"] if t["status"] == "PASS")
    failed = sum(1 for t in results["tests"] if t["status"] == "FAIL")
    skipped = sum(1 for t in results["tests"] if t["status"] == "SKIP")
    total = len(results["tests"])
    
    print(f"\nTotal Tests: {total}")
    print(f"✅ Passed: {passed}")
    print(f"❌ Failed: {failed}")
    print(f"⏭️  Skipped: {skipped}")
    print(f"Success Rate: {(passed/total*100):.1f}%\n")
    
    results["summary"] = {
        "total": total,
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "success_rate": f"{(passed/total*100):.1f}%"
    }
    
    # Save results to file
    with open("comprehensive_test_results.json", "w") as f:
        json.dump(results, f, indent=2)
    
    print("Results saved to: comprehensive_test_results.json\n")

if __name__ == "__main__":
    run_all_tests()
